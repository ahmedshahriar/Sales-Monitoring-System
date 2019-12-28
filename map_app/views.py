# import logging
# import traceback
from functools import reduce
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.shortcuts import render
from .utils.utils import calculate_income, calculate_income_last_month, calculate_change
from .models import Income
import openpyxl


def map(request):
    income_obj_count = Income.objects.all().count()
    # check if the record is empty
    if income_obj_count > 1:
        return render(request, 'map.html')
    else:
        messages.error(request, 'Please insert data!', extra_tags='error')
        return HttpResponseRedirect('/')
    
    
def upload(request):
    return render(request, 'upload.html')


def top_sales(request):
    income_obj_count = Income.objects.all().count()
    # check if the record is empty
    if income_obj_count > 1:
        income_obj_top_list = Income.objects.order_by('income_last_month').reverse()[:10]
    
        context = {
            'income_obj_top_list': income_obj_top_list
        }
        return render(request, 'top_sales.html', context)
    else:
        messages.error(request, 'No Data to display!', extra_tags='error')
        return HttpResponseRedirect('/')


def top_changed_sales(request):
    income_obj_count = Income.objects.all().count()
    # check if the record is empty
    if income_obj_count > 1:
        income_obj_top_changed_list = Income.objects.order_by('income_difference').reverse()[:10]
        
        context_difference = {
            'income_obj_top_changed_list': income_obj_top_changed_list
        }
        return render(request, 'top_changed_sales.html', context_difference)

    else:
        messages.error(request, 'No Data to display!', extra_tags='error')
        return HttpResponseRedirect('/')
   
    
def save_to_db(wb):
    worksheet = wb["Upazila_Underweight_Stunted_Chi"]
    # print(worksheet)
    
    excel_data = list()
    district_dict = {}
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            if cell.value is None:
                cell.value = 0
            row_data.append(str(cell.value))
        if row_data[1] is not "District Name":
            district_dict[row_data[1]] = row_data[4]
        excel_data.append(row_data)
    
    del district_dict["District Code"]  # delete the cell's header
    
    for district_code, district_name in district_dict.items():
        # income_per_month, income_last_month, income_difference = ('',), 0.0, 0.0
        # try:
        #     income_per_month = calculate_income(district_code, excel_data)
        # except Exception as e:
        #     logging.error(traceback.format_exc())
        #     print(e)
        # try:
        #     income_last_month = calculate_income(district_code, excel_data)
        # except Exception as e:
        #     logging.error(traceback.format_exc())
        #     print(e)
        #
        # try:
        #     income_difference = calculate_income(district_code, excel_data)
        # except Exception as e:
        #     logging.error(traceback.format_exc())
        #     print(e)

        income_per_month = calculate_income(district_code, excel_data)
        income_last_month = calculate_income_last_month(district_code, excel_data)
        income_difference = calculate_change(district_code, excel_data)

        income_total = reduce(lambda x, y: x + y, income_per_month)
        
        try:
            obj = Income(district_code=district_code, district_name=district_name, month=3,
                         income_by_month=income_per_month, income_last_month=income_last_month,
                         income_difference=income_difference, income=income_total)
        
        except Income.DoesNotExist:
            obj = Income(district_code=district_code, district_name=district_name, month=3,
                         income_by_month=income_per_month, income_last_month=income_last_month,
                         income_difference=income_difference, income=income_total)
        obj.save()
    
    return


def index(request):
    if "GET" == request.method:
        return render(request, 'index.html')
    else:
        try:
            Income.objects.all().delete()  # Delete all the records
        except Income.DoesNotExist:
            print("Nothing to delete")
        
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        
        print(wb.active.title)
        m1, m2, m3 = wb.active.cell(
            row=1, column=8).value, wb.active.cell(row=1, column=9).value, wb.active.cell(row=1, column=10).value
        
        if wb.active.title == "Upazila_Underweight_Stunted_Chi" \
                and m1 == "month1" and m2 == "month2" and m3 == "month3":
            print("correct wb", wb.active.cell(row=1, column=9).value)
            save_to_db(wb)
            return render(request, 'map.html')
        else:
            messages.error(request, 'An error occurred!')
            return HttpResponseRedirect('/upload')
        # getting a particular sheet by name out of many sheets
        
        
        
        
        
        # income_list = Income.objects.all()
      
        
        # context = {
        #     'income_list' : income_list
        # }
        

        
        # return render(request, 'index.html', {"excel_data": excel_data})


        """check for iteration"""
        

        # if cell in worksheet['B'] and cell.value is not None:
        #     print(cell.value, end=" ")

        # for row in worksheet.iter_rows(min_row=1, min_col=1):
        #     for cell in row:
        #         # cell_dis = worksheet['B']
        #         # print(cell.value in worksheet['B'])
        #
        #         if cell in worksheet['B'] and cell.value is not None:
        #             print(cell.value, end=" ")
        #         # if cell in
        #     print()

        # for row in worksheet.iter_cols(min_row=1, min_col=1):
        #     for cell in row:
        #         if cell['District Code'].value == 4:
        #             if cell.value is not None:
        #                 print(cell.value, end=" ")
        #             # if cell.value==4 :
        #             #     print(cell.value)
        #     print()
